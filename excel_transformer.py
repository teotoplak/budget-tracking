import os
import shutil
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


# Define the transformation function
def transform_data_revolut(df):
    df['Amount (abs)'] = df['Amount'].abs()
    df['Source'] = 'Revolut'
    df['6 JARS Category'] = ''
    df['Tag'] = ''
    df['Memo'] = ''
    # Select and reorder the columns
    transformed_df = df[[
        'Started Date',
        'Completed Date',
        'Type',
        'Description',
        'Amount',
        'Amount (abs)',
        'Currency',
        'Source',
        '6 JARS Category',
        'Tag',
        'Memo'
     ]]

    return transformed_df


def backup_file(excel_path, backup_folder):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f'backup_{timestamp}.xlsx'
    backup_path = os.path.join(backup_folder, backup_filename)
    shutil.copy(excel_path, backup_path)
    print(f"Backup created at {backup_path}")


def add_dropdown(excel_path, combined_df):
    # Reload the workbook to add the dropdown list
    wb = load_workbook(excel_path)
    ws = wb['spending']
    # Define the dropdown list values for "6 JARS Category"
    dropdown_list = ['Needs', 'Wants', 'Savings', 'Education', 'Play', 'Give']
    # Create a data validation object
    dv = DataValidation(type="list", formula1=f'"{",".join(dropdown_list)}"', allow_blank=True)

    # Add the data validation to the "6 JARS Category" column (G column, which is the 7th column)
    for row in range(2, len(combined_df) + 2):
        cell = f'I{row}'
        dv.add(ws[cell])

    # Add the data validation object to the worksheet
    ws.add_data_validation(dv)

    # Save the workbook
    wb.save(excel_path)

    print("Data merged, saved, and dropdowns added. Excel file updated at 'personal_finances_test.xlsx'.")


def transform_file(input_file, output_file, backup_folder, source):
    # Load the CSV file
    df = pd.read_csv(input_file)

    if source == 'revolut':
        transformed_df = transform_data_revolut(df)
    else:
        print("Unsupported source. Please provide a valid source.")
        return

    existing_df = pd.read_excel(output_file, sheet_name='spending')
    # Create a backup of the existing Excel file
    backup_file(output_file, backup_folder)

    # Combine the existing and new data
    combined_df = pd.concat([existing_df, transformed_df], ignore_index=True)

    # Write the combined data back to the Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        combined_df.to_excel(writer, sheet_name='spending', index=False)

    add_dropdown(output_file, combined_df)

    return transformed_df


